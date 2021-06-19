# 10.24:  Featured module:  openpyxl.

# (Note:  if not using Anaconda Python, this module must be
# installed separately.)

import openpyxl as ox


# read an Excel workbook from a file
# read_only=True will limit your actions to reading
wb = ox.load_workbook('../revenue.xlsx', read_only=True)


# show list of sheets within the workbook
print(wb.sheetnames)
  # ['transactions']


# access a single sheet within the workbook
ws = wb['transactions']


# loop through the entire worksheet and look at each row
for row in ws.iter_rows():

    # subscript out two fields from the row
    col1 = row[0].value
    col2 = row[1].value


# to skip rows (for example, headers) you can use min_row= argument:
for row in ws.iter_rows(min_row=2):

    # subscript out two fields from the row
    col1 = row[0].value
    col2 = row[1].value

# note that older versions of openpyxl may use the skip_rows= argument instead of min_row=.


# access one cell by Excel cellname
c = ws['A2']           # 'cell' object
print(c.value)          # the value in the cell

print(ws['B2'].value)   # access value in one statement


